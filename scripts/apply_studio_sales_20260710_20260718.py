import json
import shutil
from datetime import datetime
from pathlib import Path

import openpyxl


SOURCE = Path(r"C:\Users\user\Downloads\sales_20260710-20260718_7d30c328-363d-4e38-ada2-7bad0f3542b4.xlsx")
REPORT_ROOT = Path(r"D:\광고보고서")
ARCHIVE_DIR = REPORT_ROOT / "data" / "smartstore_sales_summary"
PUBLIC_DIRS = [
    REPORT_ROOT / "public_dashboard" / "data",
    Path(r"C:\Users\user\Documents\New project 4\public_dashboard\data"),
]
TARGET_FILES = ["monthly-dashboard-2026-07.json", "monthly-dashboard-latest.json"]
STORE_NAME = "스마트스토어(더도톰스튜디오)"


def read_rows():
    workbook = openpyxl.load_workbook(SOURCE, read_only=True, data_only=True)
    sheet = workbook["SALES"]
    headers = [cell.value for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    required = ["날짜", "채널", "상품결제건수", "환불건수", "판매금액(총)", "판매금액(순)", "환불금액"]
    missing = [name for name in required if name not in headers]
    if missing:
        raise ValueError(f"Missing columns: {missing}")
    index = {name: headers.index(name) for name in required}
    rows = []
    for values in sheet.iter_rows(min_row=2, values_only=True):
        date = str(values[index["날짜"]])[:10]
        if not ("2026-07-10" <= date <= "2026-07-18"):
            raise ValueError(f"Unexpected date: {date}")
        channel = str(values[index["채널"]] or "")
        if "더도톰스튜디오" not in channel:
            raise ValueError(f"Unexpected channel: {channel}")
        rows.append({
            "date": date,
            "grossSales": int(values[index["판매금액(총)"]] or 0),
            "refunds": int(values[index["환불금액"]] or 0),
            "netSales": int(values[index["판매금액(순)"]] or 0),
            "orders": int(values[index["상품결제건수"]] or 0),
            "refundOrders": int(values[index["환불건수"]] or 0),
        })
    if len(rows) != 9 or len({row["date"] for row in rows}) != 9:
        raise ValueError("Expected exactly one row for each date from July 10 through July 18")
    return rows


def week_index(date_text):
    day = int(date_text[-2:])
    if day <= 5:
        return 0
    if day <= 12:
        return 1
    if day <= 19:
        return 2
    if day <= 26:
        return 3
    return 4


def update_dashboard(path, imported_rows):
    data = json.loads(path.read_text(encoding="utf-8-sig"))
    metrics = data["totalSalesMetrics"]
    daily_by_date = {item["date"]: item for item in metrics.get("dailySales", [])}
    source_name = SOURCE.name

    for row in imported_rows:
        day = daily_by_date.setdefault(row["date"], {"date": row["date"], "weekday": "", "stores": []})
        stores = [store for store in day.get("stores", []) if store.get("name") != STORE_NAME]
        stores.append({
            "name": STORE_NAME,
            "store": STORE_NAME,
            "grossSales": row["grossSales"],
            "refunds": row["refunds"],
            "netSales": row["netSales"],
            "orders": row["orders"],
            "refundOrders": row["refundOrders"],
            "source": source_name,
            "basis": "스마트스토어 판매성과 전체 행 기준 판매금액(순)",
        })
        day["stores"] = stores
        day["totalGrossSales"] = sum(int(store.get("grossSales", 0)) for store in stores)
        day["totalRefunds"] = sum(int(store.get("refunds", 0)) for store in stores)
        day["totalNetSales"] = sum(int(store.get("netSales", 0)) for store in stores)
        day["netSales"] = day["totalNetSales"]

    metrics["dailySales"] = [daily_by_date[key] for key in sorted(daily_by_date)]
    metrics["dailySalesTotals"] = [
        {"date": day["date"], "netSales": day["totalNetSales"]}
        for day in metrics["dailySales"]
    ]
    metrics["dailySalesUpdatedAt"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    studio_weeks = [0, 0, 0, 0, 0]
    for day in metrics["dailySales"]:
        for store in day.get("stores", []):
            if store.get("name") == STORE_NAME:
                studio_weeks[week_index(day["date"])] += int(store.get("netSales", 0))

    auto_sales = metrics.get("autoSales", [])
    studio = next((item for item in auto_sales if item.get("name") == "더도톰스튜디오"), None)
    if studio is None:
        studio = {"name": "더도톰스튜디오", "weeks": [0, 0, 0, 0, 0], "total": 0}
        auto_sales.append(studio)
    studio["weeks"] = studio_weeks
    studio["total"] = sum(studio_weeks)
    metrics["autoSales"] = auto_sales

    combined_weeks = [0, 0, 0, 0, 0]
    for item in auto_sales:
        for i, value in enumerate(item.get("weeks", [])[:5]):
            combined_weeks[i] += float(value or 0)
    metrics["autoSalesTotal"] = {"name": "자동 매출 합계", "weeks": combined_weeks, "total": sum(combined_weeks)}

    revenue = data.setdefault("revenueMetrics", {})
    revenue["dailySales"] = metrics["dailySales"]
    revenue["dailySalesUpdatedAt"] = metrics["dailySalesUpdatedAt"]
    revenue["updatedAt"] = metrics["dailySalesUpdatedAt"]
    path.write_text(json.dumps(data, ensure_ascii=False, separators=(",", ":")), encoding="utf-8")


def main():
    imported_rows = read_rows()
    ARCHIVE_DIR.mkdir(parents=True, exist_ok=True)
    archived = ARCHIVE_DIR / SOURCE.name
    shutil.copy2(SOURCE, archived)
    updated = []
    for directory in PUBLIC_DIRS:
        for filename in TARGET_FILES:
            path = directory / filename
            update_dashboard(path, imported_rows)
            updated.append(str(path))
    print(json.dumps({
        "source": str(archived),
        "dates": [row["date"] for row in imported_rows],
        "grossSales": sum(row["grossSales"] for row in imported_rows),
        "refunds": sum(row["refunds"] for row in imported_rows),
        "netSales": sum(row["netSales"] for row in imported_rows),
        "orders": sum(row["orders"] for row in imported_rows),
        "refundOrders": sum(row["refundOrders"] for row in imported_rows),
        "updated": updated,
    }, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()


