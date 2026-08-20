from __future__ import annotations
import json
from collections import defaultdict
from datetime import datetime
from pathlib import Path
import openpyxl

ROOT=Path(__file__).resolve().parents[1]
PRODUCTS={"수건단품":"4624494637","세트모음전":"4843121925","조구만고리수건":"12924495111"}

def num(v):
    return float(v or 0) if isinstance(v,(int,float)) else 0.0

def main(source: Path):
    wb=openpyxl.load_workbook(source,data_only=True,read_only=True)
    perf=[]
    for ws in [s for s in wb.worksheets if s.title.startswith("제품 성과(") and s.title != "제품 성과(9월)"]:
        current=None
        for row in ws.iter_rows(min_row=2,max_col=10,values_only=True):
            product,date,ad_sales,ad_cost,status,sales,orders=row[:7]
            if isinstance(date,datetime): current=date.date()
            if product in PRODUCTS and current and any(v is not None for v in (ad_sales,ad_cost,sales,orders)):
                perf.append({"product":product,"date":str(current),"adSales":round(num(ad_sales)),"adCost":round(num(ad_cost)),"sales":round(num(sales)),"orders":round(num(orders)),"status":str(status or "")})
    ranks=defaultdict(list)
    ws=wb["순위 트래킹"]
    headers=next(ws.iter_rows(min_row=1,max_row=1,max_col=80,values_only=True))
    for offset,value in enumerate(headers,1):
        if value != "날짜&상품": continue
        current=None
        for product,date,keyword,rank in ws.iter_rows(min_row=2,min_col=offset,max_col=offset+3,values_only=True):
            if isinstance(date,datetime): current=date.date()
            if product in PRODUCTS and current and isinstance(rank,(int,float)):
                ranks[(product,str(current))].append(float(rank))
    slot_rows=[]
    ws=wb["슬론(리워드)관리"]
    for row in ws.iter_rows(min_row=2,max_col=8,values_only=True):
        product,start,end,cost,vendor,status,keyword,link=row
        if product in PRODUCTS and isinstance(start,datetime) and isinstance(end,datetime):
            slot_rows.append({"product":product,"startDate":str(start.date()),"endDate":str(end.date()),"cost":round(num(cost)),"status":str(status or "")})
    output=[]
    for product,pid in PRODUCTS.items():
        rows=sorted((r for r in perf if r["product"]==product),key=lambda r:r["date"])
        months=[]
        for month in sorted({r["date"][:7] for r in rows}):
            mr=[r for r in rows if r["date"].startswith(month)]
            sales=sum(r["sales"] for r in mr); orders=sum(r["orders"] for r in mr); days=len({r["date"] for r in mr})
            months.append({"month":month,"days":days,"sales":sales,"orders":orders,"dailySales":round(sales/days) if days else 0,"dailyOrders":orders/days if days else 0,"orderValue":round(sales/orders) if orders else 0,"adCost":sum(r["adCost"] for r in mr),"adSales":sum(r["adSales"] for r in mr)})
        rank_daily=[{"date":date,"rank":sum(values)/len(values),"samples":len(values)} for (p,date),values in sorted(ranks.items(),key=lambda x:x[0][1]) if p==product]
        slots=[r for r in slot_rows if r["product"]==product]
        latest=max((r["date"] for r in rows),default="")
        aug=next((m for m in months if m["month"]=="2026-08"),{})
        output.append({"product":product,"productId":pid,"latestDate":latest,"current":aug,"months":months,"rankDaily":rank_daily,"slotStart":min((r["startDate"] for r in slots),default=""),"slotCost":sum(r["cost"] for r in slots),"slotPeriods":slots})
    payload={"source":"Google Sheets 제품 성과·순위 트래킹·슬론(리워드)관리","sourceUrl":"https://docs.google.com/spreadsheets/d/1JEW2j1kRDo5P0sIEJQxXlGgk9Ao3NaHc_6B9eMhBAxM/edit","updatedAt":datetime.now().strftime("%Y-%m-%d %H:%M:%S"),"products":output}
    target=ROOT/"data"/"slot-product-overview-2026-08.json"
    target.write_text(json.dumps(payload,ensure_ascii=False,separators=(",",":")),encoding="utf-8")
    print(target)

if __name__=="__main__":
    import sys
    main(Path(sys.argv[1]))