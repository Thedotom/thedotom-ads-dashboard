from __future__ import annotations

import argparse
import json
import re
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
DATA = ROOT / "data"
PRODUCT_IDS = {
    "수건단품": "4624494637",
    "세트모음전": "4843121925",
    "조구만고리수건": "12924495111",
}


def as_number(value) -> float:
    return float(value or 0)


def load_source(path: Path) -> tuple[list[dict], dict]:
    workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
    performance: dict[tuple, dict] = {}
    for month in range(1, 8):
        sheet = workbook[f"제품 성과({month}월)"]
        for row in sheet.iter_rows(min_row=2, max_col=10, values_only=True):
            product, date_value, ad_sales, ad_cost, status, sales, orders = row[:7]
            if not product or not isinstance(date_value, datetime):
                continue
            performance[(date_value.date(), str(product))] = {
                "adSales": as_number(ad_sales),
                "adCost": as_number(ad_cost),
                "status": str(status or ""),
                "sales": as_number(sales),
                "orders": as_number(orders),
            }

    rank_rows: dict[tuple, list[float]] = defaultdict(list)
    rank_sheet = workbook["순위 트래킹"]
    for offset in (1, 6, 11, 16):
        current_date = None
        for row in rank_sheet.iter_rows(
            min_row=2, min_col=offset, max_col=offset + 3, values_only=True
        ):
            product, date_value, _keyword, rank = row
            if isinstance(date_value, datetime):
                current_date = date_value.date()
            if product and current_date and isinstance(rank, (int, float)):
                rank_rows[(current_date, str(product))].append(float(rank))

    def summarize(product: str, start, end) -> dict:
        values = []
        day = start
        while day <= end:
            value = performance.get((day, product))
            if value is not None:
                values.append(value)
            day += timedelta(days=1)
        return {
            "daysWithData": len(values),
            "sales": round(sum(value["sales"] for value in values)),
            "orders": round(sum(value["orders"] for value in values)),
            "adSales": round(sum(value["adSales"] for value in values)),
            "adCost": round(sum(value["adCost"] for value in values)),
        }

    periods = []
    slot_sheet = workbook["슬론(리워드)관리"]
    for row in slot_sheet.iter_rows(min_row=2, max_col=8, values_only=True):
        product, start_value, end_value, slot_cost, vendor, source_status, keyword, link = row
        if not product or not isinstance(start_value, datetime) or not isinstance(end_value, datetime):
            continue
        product = str(product)
        start = start_value.date()
        end = end_value.date()
        days = (end - start).days + 1
        prior_end = start - timedelta(days=1)
        prior_start = prior_end - timedelta(days=days - 1)
        during = summarize(product, start, end)
        prior = summarize(product, prior_start, prior_end)
        cost = round(as_number(slot_cost))
        incremental_sales = during["sales"] - prior["sales"]
        rank_values = []
        day = start
        while day <= end:
            rank_values.extend(rank_rows.get((day, product), []))
            day += timedelta(days=1)
        complete = during["daysWithData"] >= days
        sales_change_rate = incremental_sales / prior["sales"] if prior["sales"] else None
        slot_cost_rate = cost / during["sales"] if during["sales"] else None
        blended_cost = during["adCost"] + cost
        blended_roas = during["sales"] / blended_cost if blended_cost else None
        average_rank = sum(rank_values) / len(rank_values) if rank_values else None
        if not complete:
            decision = "관찰중"
        elif average_rank is not None and average_rank <= 5 and (
            slot_cost_rate is None or slot_cost_rate <= 0.05
        ):
            decision = "효율 양호"
        elif average_rank is not None and average_rank <= 5:
            decision = "순위 유지"
        elif slot_cost_rate is not None and slot_cost_rate > 0.15:
            decision = "비용 주의"
        else:
            decision = "재검토"
        periods.append(
            {
                "product": product,
                "productId": PRODUCT_IDS.get(product, ""),
                "keyword": str(keyword or ""),
                "link": str(link or ""),
                "vendor": str(vendor or ""),
                "startDate": str(start),
                "endDate": str(end),
                "comparisonStatus": "pending_2025_data",
                "comparisonBasis": "전년 동기간 매출",
                "days": days,
                "sourceStatus": str(source_status or ""),
                "isComplete": complete,
                "slotCost": cost,
                "yearOverYearSales": None,
                "during": during,
                "incrementalSales": None,
                "salesChangeRate": None,
                "incrementalRoas": None,
                "slotCostRate": slot_cost_rate,
                "blendedMarketingCost": round(blended_cost),
                "blendedSalesRoas": blended_roas,
                "rankAverage": average_rank,
                "rankBest": min(rank_values) if rank_values else None,
                "rankWorst": max(rank_values) if rank_values else None,
                "rankSamples": len(rank_values),
                "decision": decision,
            }
        )

    completed = [period for period in periods if period["isComplete"]]
    summary = {
        "periodCount": len(periods),
        "completedCount": len(completed),
        "activeCount": len(periods) - len(completed),
        "totalSlotCost": sum(period["slotCost"] for period in completed),
        "totalSales": sum(period["during"]["sales"] for period in completed),
        "totalPriorSales": None,
        "incrementalSales": None,
        "totalAdCost": sum(period["during"]["adCost"] for period in completed),
    }
    summary["slotCostRate"] = (
        summary["totalSlotCost"] / summary["totalSales"] if summary["totalSales"] else 0
    )
    total_marketing_cost = summary["totalSlotCost"] + summary["totalAdCost"]
    summary["blendedSalesRoas"] = (
        summary["totalSales"] / total_marketing_cost if total_marketing_cost else 0
    )
    ranks = [
        period["rankAverage"]
        for period in completed
        if period["rankAverage"] is not None
    ]
    summary["rankAverage"] = sum(ranks) / len(ranks) if ranks else None

    products = []
    for product in PRODUCT_IDS:
        rows = [period for period in completed if period["product"] == product]
        if not rows:
            continue
        sales = sum(row["during"]["sales"] for row in rows)
        incremental = None
        product_ranks = [
            row["rankAverage"] for row in rows if row["rankAverage"] is not None
        ]
        products.append(
            {
                "product": product,
                "productId": PRODUCT_IDS[product],
                "periodCount": len(rows),
                "slotCost": sum(row["slotCost"] for row in rows),
                "sales": sales,
                "incrementalSales": incremental,
                "slotCostRate": (
                    sum(row["slotCost"] for row in rows) / sales if sales else 0
                ),
                "rankAverage": (
                    sum(product_ranks) / len(product_ranks) if product_ranks else None
                ),
                "bestRank": min(
                    row["rankBest"] for row in rows if row["rankBest"] is not None
                ),
                "goodPeriods": sum(row["decision"] == "효율 우수" for row in rows),
                "defensePeriods": sum(row["decision"] == "순위 방어" for row in rows),
            }
        )

    return periods, {
        "source": path.name,
        "sourceUrl": "https://docs.google.com/spreadsheets/d/1JEW2j1kRDo5P0sIEJQxXlGgk9Ao3NaHc_6B9eMhBAxM/edit",
        "calculationBasis": "슬롯 운영 기간의 매출·주문·광고비·검색 순위 집계",
        "comparisonStatus": "2025년 전년 동기간 매출 데이터 적재 대기",
        "updatedAt": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "summary": summary,
        "products": products,
        "periods": periods,
    }


def update_dashboard_files(payload: dict) -> None:
    targets = sorted(path for path in DATA.glob("monthly-dashboard-2026-*.json") if re.fullmatch(r"monthly-dashboard-2026-\d{2}\.json", path.name))
    latest = DATA / "monthly-dashboard-latest.json"
    if latest.exists():
        targets.append(latest)
    for path in targets:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        data["slotEfficiency"] = payload
        path.write_text(
            json.dumps(data, ensure_ascii=False, separators=(",", ":")),
            encoding="utf-8",
        )
        print(f"updated {path.name}")


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("source", type=Path)
    args = parser.parse_args()
    periods, payload = load_source(args.source)
    update_dashboard_files(payload)
    print(
        json.dumps(
            {
                "periods": len(periods),
                "completed": payload["summary"]["completedCount"],
                "products": len(payload["products"]),
            },
            ensure_ascii=False,
        )
    )


if __name__ == "__main__":
    main()

